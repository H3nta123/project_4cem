"""
ФинансПро — FastAPI application v2.1.0
All endpoints for budget, categories, scenarios, transactions,
investments, savings, loans, settings, import, export, analytics.
"""

import csv
import io
import re
import json
import logging
import os
import sys
from datetime import datetime, timedelta

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy.orm import Session

from database import (
    get_db, init_db, SessionLocal,
    BudgetCategory, Scenario, Transaction, Investment, SavingsGoal, Loan, Setting, Base, engine,
)
from schemas import (
    CategoryCreate,
    CategoryUpdate,
    CategoryResponse,
    BudgetValueUpdate,
    BudgetFactValueUpdate,
    BulkBudgetUpdate,
    MonthlyUpdate,
    AutofillRequest,
    ScenarioCreate,
    ScenarioResponse,
    BudgetTableResponse,
    TransactionCreate,
    TransactionResponse,
    InvestmentCreate,
    InvestmentUpdate,
    InvestmentResponse,
    SavingsGoalCreate,
    SavingsGoalUpdate,
    SavingsGoalResponse,
    LoanCreate,
    LoanUpdate,
    LoanResponse,
    SettingsResponse,
    SettingsUpdate,
    AnalyticsSummaryResponse,
    CategoryBreakdownItem,
    ImportResponse,
)
from seed import seed_data

# ─── Logging ─────────────────────────────────────────────
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("finanspro")

# ─── App ─────────────────────────────────────────────────
app = FastAPI(title="ФинансПро API", version="2.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        "http://localhost:5175",
        "http://127.0.0.1:5175",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Static files (bundled frontend) ─────────────────────
def _get_dist_dir():
    """Resolve the frontend dist directory."""
    if getattr(sys, 'frozen', False):
        base = sys._MEIPASS
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, 'dist')


DIST_DIR = _get_dist_dir()


@app.on_event("startup")
def on_startup():
    logger.info("Initializing database...")
    init_db()
    seed_data()
    logger.info("Database ready.")
    logger.info(f"Static files dir: {DIST_DIR}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CATEGORIES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.get("/api/categories", response_model=list[CategoryResponse])
def get_categories(db: Session = Depends(get_db)):
    """Get all budget categories ordered by sort_order."""
    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    return [_cat_to_dict(c) for c in cats]


@app.post("/api/categories", response_model=CategoryResponse, status_code=201)
def create_category(data: CategoryCreate, db: Session = Depends(get_db)):
    """Create a new budget category."""
    cat = BudgetCategory(
        name=data.name,
        type=data.type.value,
        monthly=json.dumps(data.monthly),
        monthly_fact=json.dumps(data.monthly_fact),
        sort_order=data.sort_order,
        parent_id=data.parent_id,
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    logger.info(f"Created category: {cat.name} ({cat.type})")
    return _cat_to_dict(cat)


@app.put("/api/categories/{cat_id}", response_model=CategoryResponse)
def update_category(cat_id: int, data: CategoryUpdate, db: Session = Depends(get_db)):
    """Update a category's name, type, or monthly values."""
    cat = db.query(BudgetCategory).filter(BudgetCategory.id == cat_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Категория не найдена")

    if data.name is not None:
        cat.name = data.name
    if data.type is not None:
        cat.type = data.type.value
    if data.monthly is not None:
        cat.monthly = json.dumps(data.monthly)
    if data.monthly_fact is not None:
        cat.monthly_fact = json.dumps(data.monthly_fact)
    if data.sort_order is not None:
        cat.sort_order = data.sort_order

    db.commit()
    db.refresh(cat)
    return _cat_to_dict(cat)


@app.delete("/api/categories/{cat_id}", status_code=204)
def delete_category(cat_id: int, db: Session = Depends(get_db)):
    """Delete a category by ID."""
    cat = db.query(BudgetCategory).filter(BudgetCategory.id == cat_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Категория не найдена")
    db.delete(cat)
    db.commit()
    logger.info(f"Deleted category id={cat_id}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  BUDGET TABLE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.get("/api/budget", response_model=BudgetTableResponse)
def get_budget_table(db: Session = Depends(get_db)):
    """Get the full budget table with totals."""
    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    return _build_budget_response(cats, db=db)


@app.put("/api/budget/values/{cat_id}", response_model=CategoryResponse)
def update_budget_values(cat_id: int, data: MonthlyUpdate, db: Session = Depends(get_db)):
    """Replace all monthly values for a category."""
    cat = db.query(BudgetCategory).filter(BudgetCategory.id == cat_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Категория не найдена")
    cat.monthly = json.dumps(data.monthly)
    
    # Duplicate to fact
    monthly_fact = json.loads(cat.monthly_fact) if cat.monthly_fact else []
    while len(monthly_fact) < len(data.monthly):
        monthly_fact.append(0)
    for i, val in enumerate(data.monthly):
        if i < len(monthly_fact):
            monthly_fact[i] = val
    cat.monthly_fact = json.dumps(monthly_fact)
    
    db.commit()
    db.refresh(cat)
    return _cat_to_dict(cat)


@app.put("/api/budget/cell", response_model=BudgetTableResponse)
def update_single_cell(data: BudgetValueUpdate, db: Session = Depends(get_db)):
    """Update a single cell in the budget table (plan) and return refreshed totals."""
    cat = db.query(BudgetCategory).filter(BudgetCategory.id == data.category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Категория не найдена")

    monthly = json.loads(cat.monthly)
    # Extend if needed
    while len(monthly) < 52:
        monthly.append(0)
    if not (0 <= data.month_index < len(monthly)):
        raise HTTPException(status_code=400, detail="Индекс месяца вне диапазона")

    monthly[data.month_index] = data.value
    cat.monthly = json.dumps(monthly)
    
    monthly_fact = json.loads(cat.monthly_fact)
    while len(monthly_fact) < 52:
        monthly_fact.append(0)
    monthly_fact[data.month_index] = data.value
    cat.monthly_fact = json.dumps(monthly_fact)
    
    db.commit()

    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    return _build_budget_response(cats, db=db)


@app.put("/api/budget/fact/cell", response_model=BudgetTableResponse)
def update_fact_cell(data: BudgetFactValueUpdate, db: Session = Depends(get_db)):
    """Update a single cell in the fact table and return refreshed totals."""
    cat = db.query(BudgetCategory).filter(BudgetCategory.id == data.category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Категория не найдена")

    monthly_fact = json.loads(cat.monthly_fact)
    while len(monthly_fact) < 52:
        monthly_fact.append(0)
    if not (0 <= data.month_index < len(monthly_fact)):
        raise HTTPException(status_code=400, detail="Индекс недели вне диапазона")

    monthly_fact[data.month_index] = data.value
    cat.monthly_fact = json.dumps(monthly_fact)
    db.commit()

    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    return _build_budget_response(cats, db=db)


@app.put("/api/budget/bulk", response_model=BudgetTableResponse)
def bulk_update_budget(data: BulkBudgetUpdate, db: Session = Depends(get_db)):
    """Batch-update multiple cells at once."""
    for upd in data.updates:
        cat = db.query(BudgetCategory).filter(BudgetCategory.id == upd.category_id).first()
        if not cat:
            continue
        monthly = json.loads(cat.monthly)
        while len(monthly) < 52:
            monthly.append(0)
        if 0 <= upd.month_index < len(monthly):
            monthly[upd.month_index] = upd.value
            cat.monthly = json.dumps(monthly)
            
            monthly_fact = json.loads(cat.monthly_fact)
            while len(monthly_fact) < 52:
                monthly_fact.append(0)
            monthly_fact[upd.month_index] = upd.value
            cat.monthly_fact = json.dumps(monthly_fact)
    db.commit()

    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    return _build_budget_response(cats, db=db)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  AUTOFILL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.put("/api/budget/autofill", response_model=BudgetTableResponse)
def autofill_budget(data: AutofillRequest, db: Session = Depends(get_db)):
    """Autofill budget values for a category.
    period: 'weekly' fills each week, 'monthly' fills every 4 weeks.
    count: number of periods to fill.
    target: 'plan' or 'fact'.
    """
    cat = db.query(BudgetCategory).filter(BudgetCategory.id == data.category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Категория не найдена")

    if data.target == "fact":
        values = json.loads(cat.monthly_fact)
    else:
        values = json.loads(cat.monthly)

    while len(values) < 52:
        values.append(0)

    step = 1 if data.period == "weekly" else 4  # monthly = every 4 weeks
    week = data.start_week
    for _ in range(data.count):
        if 0 <= week < 52:
            values[week] = data.amount
        week += step

    if data.target == "fact":
        cat.monthly_fact = json.dumps(values)
    else:
        cat.monthly = json.dumps(values)
        
        # Duplicate to fact
        monthly_fact = json.loads(cat.monthly_fact) if cat.monthly_fact else []
        while len(monthly_fact) < 52:
            monthly_fact.append(0)
        
        week = data.start_week
        for _ in range(data.count):
            if 0 <= week < 52:
                monthly_fact[week] = data.amount
            week += step
        cat.monthly_fact = json.dumps(monthly_fact)

    db.commit()

    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    return _build_budget_response(cats, db=db)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SCENARIOS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.get("/api/scenarios", response_model=list[ScenarioResponse])
def get_scenarios(db: Session = Depends(get_db)):
    """Get all available scenarios."""
    return db.query(Scenario).all()


@app.post("/api/scenarios", response_model=ScenarioResponse, status_code=201)
def create_scenario(data: ScenarioCreate, db: Session = Depends(get_db)):
    """Create a custom scenario."""
    scenario = Scenario(**data.model_dump())
    db.add(scenario)
    db.commit()
    db.refresh(scenario)
    return scenario


@app.delete("/api/scenarios/{scenario_id}", status_code=204)
def delete_scenario(scenario_id: int, db: Session = Depends(get_db)):
    """Delete a custom scenario."""
    scenario = db.query(Scenario).filter(Scenario.id == scenario_id).first()
    if not scenario:
        raise HTTPException(status_code=404, detail="Сценарий не найден")
    if scenario.type != "custom":
        raise HTTPException(status_code=400, detail="Нельзя удалить встроенный сценарий")
    db.delete(scenario)
    db.commit()


@app.get("/api/budget/scenario/{scenario_id}", response_model=BudgetTableResponse)
def get_budget_with_scenario(scenario_id: int, db: Session = Depends(get_db)):
    """Get budget table with a scenario applied (values modified by modifiers)."""
    scenario = db.query(Scenario).filter(Scenario.id == scenario_id).first()
    if not scenario:
        raise HTTPException(status_code=404, detail="Сценарий не найден")
    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    return _build_budget_response(cats, scenario, db=db)


@app.post("/api/scenarios/{scenario_id}/apply", response_model=BudgetTableResponse)
def apply_scenario_to_plan(scenario_id: int, db: Session = Depends(get_db)):
    """Apply a scenario to the main plan — copies modified values into monthly (plan)."""
    scenario = db.query(Scenario).filter(Scenario.id == scenario_id).first()
    if not scenario:
        raise HTTPException(status_code=404, detail="Сценарий не найден")

    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    for cat in cats:
        monthly = json.loads(cat.monthly)
        while len(monthly) < 52:
            monthly.append(0)

        if cat.type == "income":
            monthly = [round(v * scenario.income_modifier, 2) for v in monthly]
        else:
            per_week_extra = scenario.extra_expense / 52
            monthly = [round(v * scenario.expense_modifier + per_week_extra, 2) for v in monthly]

        cat.monthly = json.dumps(monthly)

    db.commit()
    logger.info(f"Applied scenario '{scenario.name}' to plan")

    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    return _build_budget_response(cats, db=db)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  TRANSACTIONS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.get("/api/transactions", response_model=list[TransactionResponse])
def get_transactions(
    type: str | None = Query(None, description="Filter by type: income or expense"),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    """Get transactions with optional filtering and pagination."""
    q = db.query(Transaction)
    if type:
        q = q.filter(Transaction.type == type)
    q = q.order_by(Transaction.date.desc(), Transaction.id.desc())
    return q.offset(offset).limit(limit).all()


@app.post("/api/transactions", response_model=TransactionResponse, status_code=201)
def create_transaction(data: TransactionCreate, db: Session = Depends(get_db)):
    """Create a new transaction."""
    tx = Transaction(
        date=data.date,
        category=data.category,
        description=data.description,
        amount=data.amount,
        type=data.type.value,
    )
    db.add(tx)
    db.commit()
    db.refresh(tx)
    logger.info(f"Created transaction: {tx.category} {tx.amount} ({tx.type})")
    return tx


@app.delete("/api/transactions/{tx_id}", status_code=204)
def delete_transaction(tx_id: int, db: Session = Depends(get_db)):
    """Delete a transaction by ID."""
    tx = db.query(Transaction).filter(Transaction.id == tx_id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Транзакция не найдена")
    db.delete(tx)
    db.commit()
    logger.info(f"Deleted transaction id={tx_id}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  INVESTMENTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.get("/api/investments", response_model=list[InvestmentResponse])
def get_investments(db: Session = Depends(get_db)):
    """Get all investments."""
    return db.query(Investment).order_by(Investment.id.desc()).all()


@app.post("/api/investments", response_model=InvestmentResponse, status_code=201)
def create_investment(data: InvestmentCreate, db: Session = Depends(get_db)):
    """Create a new investment."""
    inv = Investment(
        name=data.name,
        type=data.type.value,
        type_label=data.type_label,
        value=data.value,
        purchase_price=data.purchase_price,
        growth=data.growth,
        quantity=data.quantity,
    )
    db.add(inv)
    db.commit()
    db.refresh(inv)
    logger.info(f"Created investment: {inv.name}")
    return inv


@app.put("/api/investments/{inv_id}", response_model=InvestmentResponse)
def update_investment(inv_id: int, data: InvestmentUpdate, db: Session = Depends(get_db)):
    """Update an investment (e.g. after buy/sell)."""
    inv = db.query(Investment).filter(Investment.id == inv_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Инвестиция не найдена")

    if data.name is not None:
        inv.name = data.name
    if data.value is not None:
        inv.value = data.value
    if data.purchase_price is not None:
        inv.purchase_price = data.purchase_price
    if data.growth is not None:
        inv.growth = data.growth
    if data.quantity is not None:
        inv.quantity = data.quantity

    db.commit()
    db.refresh(inv)
    return inv


@app.delete("/api/investments/{inv_id}", status_code=204)
def delete_investment(inv_id: int, db: Session = Depends(get_db)):
    """Delete an investment."""
    inv = db.query(Investment).filter(Investment.id == inv_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Инвестиция не найдена")
    db.delete(inv)
    db.commit()
    logger.info(f"Deleted investment id={inv_id}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SAVINGS GOALS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.get("/api/savings", response_model=list[SavingsGoalResponse])
def get_savings(db: Session = Depends(get_db)):
    """Get all savings goals."""
    return db.query(SavingsGoal).order_by(SavingsGoal.id.desc()).all()


@app.post("/api/savings", response_model=SavingsGoalResponse, status_code=201)
def create_saving(data: SavingsGoalCreate, db: Session = Depends(get_db)):
    """Create a new savings goal."""
    cat = BudgetCategory(
        name=f"Копилка: {data.name}",
        type="expense",
        monthly="[" + ",".join(["0.0"]*52) + "]",
        monthly_fact="[" + ",".join(["0.0"]*52) + "]",
        sort_order=9999
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)

    goal = SavingsGoal(
        name=data.name,
        target=data.target,
        current=data.current,
        category_id=cat.id,
    )
    db.add(goal)
    db.commit()
    db.refresh(goal)
    logger.info(f"Created savings goal: {goal.name} with category_id: {cat.id}")
    return goal


@app.put("/api/savings/{goal_id}", response_model=SavingsGoalResponse)
def update_saving(goal_id: int, data: SavingsGoalUpdate, db: Session = Depends(get_db)):
    """Update a savings goal (e.g. top up)."""
    goal = db.query(SavingsGoal).filter(SavingsGoal.id == goal_id).first()
    if not goal:
        raise HTTPException(status_code=404, detail="Цель не найдена")

    if data.name is not None:
        goal.name = data.name
    if data.current is not None:
        goal.current = data.current
    if data.target is not None:
        goal.target = data.target

    db.commit()
    db.refresh(goal)
    return goal


@app.delete("/api/savings/{goal_id}", status_code=204)
def delete_saving(goal_id: int, db: Session = Depends(get_db)):
    """Delete a savings goal."""
    goal = db.query(SavingsGoal).filter(SavingsGoal.id == goal_id).first()
    if not goal:
        raise HTTPException(status_code=404, detail="Цель не найдена")
    if goal.category_id:
        cat = db.query(BudgetCategory).filter(BudgetCategory.id == goal.category_id).first()
        if cat:
            db.delete(cat)
    db.delete(goal)
    db.commit()
    logger.info(f"Deleted savings goal id={goal_id}")

from schemas import SavingsGoalTopUp

@app.post("/api/savings/{goal_id}/topup", response_model=SavingsGoalResponse)
def topup_saving(goal_id: int, data: SavingsGoalTopUp, db: Session = Depends(get_db)):
    """Top up a savings goal and record the expense in the budget."""
    goal = db.query(SavingsGoal).filter(SavingsGoal.id == goal_id).first()
    if not goal:
        raise HTTPException(status_code=404, detail="Цель не найдена")

    goal.current += data.amount

    if goal.category_id:
        cat = db.query(BudgetCategory).filter(BudgetCategory.id == goal.category_id).first()
        if cat:
            curr_idx = _get_current_week_index(db)
            fact_plan = json.loads(cat.monthly_fact) if cat.monthly_fact else [0.0]*52
            while len(fact_plan) < 52: fact_plan.append(0.0)
            fact_plan[curr_idx] += data.amount
            cat.monthly_fact = json.dumps(fact_plan)

    db.commit()
    db.refresh(goal)
    return goal

@app.post("/api/savings/{goal_id}/withdraw", response_model=SavingsGoalResponse)
def withdraw_saving(goal_id: int, data: SavingsGoalTopUp, db: Session = Depends(get_db)):
    """Withdraw from a savings goal and record the negative expense in the budget."""
    goal = db.query(SavingsGoal).filter(SavingsGoal.id == goal_id).first()
    if not goal:
        raise HTTPException(status_code=404, detail="Цель не найдена")

    if data.amount > goal.current:
        raise HTTPException(status_code=400, detail="Недостаточно средств в копилке")

    goal.current -= data.amount

    if goal.category_id:
        cat = db.query(BudgetCategory).filter(BudgetCategory.id == goal.category_id).first()
        if cat:
            curr_idx = _get_current_week_index(db)
            fact_plan = json.loads(cat.monthly_fact) if cat.monthly_fact else [0.0]*52
            while len(fact_plan) < 52: fact_plan.append(0.0)
            fact_plan[curr_idx] -= data.amount
            cat.monthly_fact = json.dumps(fact_plan)

    db.commit()
    db.refresh(goal)
    return goal


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  LOANS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _generate_loan_plan(total: float, payment: float, payment_type: str, start_index: int = 0) -> list[float]:
    plan = [0.0] * 52
    if payment <= 0 or total <= 0: return plan
    remaining = total
    if payment_type == "weekly":
        for i in range(start_index, 52):
            if remaining <= 0: break
            pay = min(payment, remaining)
            plan[i] = pay
            remaining -= pay
    else:
        for i in range(0, 52, 4):
            if i < start_index: continue
            if remaining <= 0: break
            pay = min(payment, remaining)
            plan[i] = pay
            remaining -= pay
    return plan

def _get_current_week_index(db: Session) -> int:
    dates = _get_weekly_dates(db)
    today = datetime.now().strftime("%Y-%m-%d")
    idx = 0
    for i, d in enumerate(dates):
        if today >= d:
            idx = i
        else:
            break
    return idx

@app.get("/api/loans", response_model=list[LoanResponse])
def get_loans(db: Session = Depends(get_db)):
    """Get all loans."""
    return db.query(Loan).order_by(Loan.id.desc()).all()


@app.post("/api/loans", response_model=LoanResponse, status_code=201)
def create_loan(data: LoanCreate, db: Session = Depends(get_db)):
    """Create a new loan and its budget category."""
    curr_idx = _get_current_week_index(db)
    plan = _generate_loan_plan(data.total_amount, data.monthly_payment, data.payment_type, curr_idx)
    
    cat = BudgetCategory(
        name=f"Кредит: {data.name}",
        type="expense",
        monthly=json.dumps(plan),
        monthly_fact="[" + ",".join(["0.0"]*52) + "]",
        sort_order=9999
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)

    loan = Loan(
        name=data.name,
        total_amount=data.total_amount,
        paid_amount=data.paid_amount,
        monthly_payment=data.monthly_payment,
        interest_rate=data.interest_rate,
        start_date=data.start_date,
        end_date=data.end_date,
        status=data.status.value,
        payment_type=data.payment_type,
        category_id=cat.id,
    )
    db.add(loan)
    db.commit()
    db.refresh(loan)
    logger.info(f"Created loan: {loan.name} with category_id: {cat.id}")
    return loan


@app.put("/api/loans/{loan_id}", response_model=LoanResponse)
def update_loan(loan_id: int, data: LoanUpdate, db: Session = Depends(get_db)):
    """Update a loan (e.g. make a payment)."""
    loan = db.query(Loan).filter(Loan.id == loan_id).first()
    if not loan:
        raise HTTPException(status_code=404, detail="Кредит не найден")

    if data.name is not None:
        loan.name = data.name
        if loan.category_id:
            cat = db.query(BudgetCategory).filter(BudgetCategory.id == loan.category_id).first()
            if cat:
                cat.name = f"Кредит: {data.name}"
    if data.paid_amount is not None:
        loan.paid_amount = data.paid_amount
    if data.monthly_payment is not None:
        loan.monthly_payment = data.monthly_payment
    if data.status is not None:
        loan.status = data.status.value
    if getattr(data, 'payment_type', None) is not None:
        loan.payment_type = data.payment_type

    # Update category plan if payment or type changed
    if data.monthly_payment is not None or getattr(data, 'payment_type', None) is not None:
        if loan.category_id:
            cat = db.query(BudgetCategory).filter(BudgetCategory.id == loan.category_id).first()
            if cat:
                new_plan = _generate_loan_plan(loan.monthly_payment, loan.payment_type)
                # only update future plan if not completed? For now update full plan.
                if loan.status == "completed":
                    curr_idx = _get_current_week_index(db)
                    plan = json.loads(cat.monthly)
                    for i in range(curr_idx, 52):
                        plan[i] = 0.0
                    cat.monthly = json.dumps(plan)
                else:
                    cat.monthly = json.dumps(new_plan)

    # Auto-complete if fully paid
    if loan.paid_amount >= loan.total_amount:
        loan.status = "completed"
        if loan.category_id:
            cat = db.query(BudgetCategory).filter(BudgetCategory.id == loan.category_id).first()
            if cat:
                curr_idx = _get_current_week_index(db)
                plan = json.loads(cat.monthly)
                for i in range(curr_idx, 52):
                    plan[i] = 0.0
                cat.monthly = json.dumps(plan)

    db.commit()
    db.refresh(loan)
    return loan


@app.delete("/api/loans/{loan_id}", status_code=204)
def delete_loan(loan_id: int, db: Session = Depends(get_db)):
    """Delete a loan."""
    loan = db.query(Loan).filter(Loan.id == loan_id).first()
    if not loan:
        raise HTTPException(status_code=404, detail="Кредит не найден")
    if loan.category_id:
        cat = db.query(BudgetCategory).filter(BudgetCategory.id == loan.category_id).first()
        if cat:
            db.delete(cat)
    db.delete(loan)
    db.commit()
    logger.info(f"Deleted loan id={loan_id}")

from schemas import LoanPay

@app.post("/api/loans/{loan_id}/pay", response_model=LoanResponse)
def pay_loan(loan_id: int, data: LoanPay, db: Session = Depends(get_db)):
    """Make a payment on a loan and record it in the budget."""
    loan = db.query(Loan).filter(Loan.id == loan_id).first()
    if not loan:
        raise HTTPException(status_code=404, detail="Кредит не найден")

    loan.paid_amount += data.amount
    if loan.paid_amount >= loan.total_amount:
        loan.status = "completed"

    if loan.category_id:
        cat = db.query(BudgetCategory).filter(BudgetCategory.id == loan.category_id).first()
        if cat:
            curr_idx = _get_current_week_index(db)
            fact_plan = json.loads(cat.monthly_fact) if cat.monthly_fact else [0.0]*52
            while len(fact_plan) < 52: fact_plan.append(0.0)
            fact_plan[curr_idx] += data.amount
            cat.monthly_fact = json.dumps(fact_plan)

            remaining = loan.total_amount - loan.paid_amount
            old_plan = json.loads(cat.monthly) if cat.monthly else [0.0]*52
            while len(old_plan) < 52: old_plan.append(0.0)
            new_plan = [0.0] * 52
            
            # Keep the past plan unmodified
            for i in range(curr_idx + 1):
                new_plan[i] = old_plan[i]
                
            # Recalculate future plan
            if remaining > 0:
                if loan.payment_type == "weekly":
                    for i in range(curr_idx + 1, 52):
                        if remaining <= 0: break
                        pay = min(loan.monthly_payment, remaining)
                        new_plan[i] = pay
                        remaining -= pay
                else:
                    for i in range(0, 52, 4):
                        if i <= curr_idx: continue
                        if remaining <= 0: break
                        pay = min(loan.monthly_payment, remaining)
                        new_plan[i] = pay
                        remaining -= pay
                        
            cat.monthly = json.dumps(new_plan)

    db.commit()
    db.refresh(loan)
    return loan


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SETTINGS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

SETTINGS_DEFAULTS = {
    "theme": "dark",
    "currency": "RUB",
    "currencySymbol": "₽",
    "notifyPayments": "true",
    "notifyBudgetExceed": "true",
    "budgetStartDate": "",
    "profitAdjustment": "0.0",
}


@app.get("/api/settings", response_model=SettingsResponse)
def get_settings(db: Session = Depends(get_db)):
    """Get all application settings."""
    settings_rows = db.query(Setting).all()
    settings_dict = {s.key: s.value for s in settings_rows}

    return SettingsResponse(
        theme=settings_dict.get("theme", SETTINGS_DEFAULTS["theme"]),
        currency=settings_dict.get("currency", SETTINGS_DEFAULTS["currency"]),
        currencySymbol=settings_dict.get("currencySymbol", SETTINGS_DEFAULTS["currencySymbol"]),
        notifyPayments=settings_dict.get("notifyPayments", SETTINGS_DEFAULTS["notifyPayments"]).lower() == "true",
        notifyBudgetExceed=settings_dict.get("notifyBudgetExceed", SETTINGS_DEFAULTS["notifyBudgetExceed"]).lower() == "true",
        budgetStartDate=settings_dict.get("budgetStartDate", SETTINGS_DEFAULTS["budgetStartDate"]),
        profitAdjustment=float(settings_dict.get("profitAdjustment", SETTINGS_DEFAULTS["profitAdjustment"])),
    )


@app.put("/api/settings", response_model=SettingsResponse)
def update_settings(data: SettingsUpdate, db: Session = Depends(get_db)):
    """Update application settings."""
    updates = data.model_dump(exclude_none=True)

    for key, value in updates.items():
        # Convert booleans to string for storage
        str_value = str(value).lower() if isinstance(value, bool) else str(value)
        existing = db.query(Setting).filter(Setting.key == key).first()
        if existing:
            existing.value = str_value
        else:
            db.add(Setting(key=key, value=str_value))

    db.commit()
    logger.info(f"Updated settings: {list(updates.keys())}")
    return get_settings(db)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ANALYTICS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.get("/api/analytics/summary", response_model=AnalyticsSummaryResponse)
def get_analytics_summary(db: Session = Depends(get_db)):
    """Get analytics summary: income/expense totals and category breakdowns."""
    # Sum from budget categories (weekly totals → month 0 as representative)
    cats = db.query(BudgetCategory).all()

    total_income = 0.0
    total_expense = 0.0
    income_items: list[dict] = []
    expense_items: list[dict] = []

    for cat in cats:
        monthly = json.loads(cat.monthly)
        total = sum(monthly)
        if cat.type == "income":
            total_income += total
            income_items.append({"name": cat.name, "amount": total})
        else:
            total_expense += total
            expense_items.append({"name": cat.name, "amount": total})

    # Also add actual transactions
    transactions = db.query(Transaction).all()
    tx_income = sum(tx.amount for tx in transactions if tx.type == "income")
    tx_expense = sum(tx.amount for tx in transactions if tx.type == "expense")
    total_income += tx_income
    total_expense += tx_expense

    # Build breakdowns with percentages
    def build_breakdown(items: list[dict], total: float) -> list[CategoryBreakdownItem]:
        result = []
        for item in sorted(items, key=lambda x: x["amount"], reverse=True):
            pct = round((item["amount"] / total) * 100, 1) if total > 0 else 0
            result.append(CategoryBreakdownItem(name=item["name"], amount=item["amount"], percentage=pct))
        return result

    return AnalyticsSummaryResponse(
        total_income=total_income,
        total_expense=total_expense,
        profit=total_income - total_expense,
        expense_breakdown=build_breakdown(expense_items, total_expense),
        income_breakdown=build_breakdown(income_items, total_income),
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  IMPORT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.post("/api/import/csv", response_model=ImportResponse)
async def import_csv(
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
):
    """Import budget table from CSV."""
    imported = 0
    skipped = 0
    errors: list[str] = []

    try:
        content = await file.read()
        for encoding in ['utf-8-sig', 'utf-8', 'cp1251']:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return ImportResponse(imported=0, skipped=0, errors=["Не удалось определить кодировку файла"])

        first_line = text.split('\n')[0]
        delimiter = ';' if ';' in first_line else ','

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)
        if not rows or len(rows) < 2:
            return ImportResponse(imported=0, skipped=0, errors=["Файл пуст или некорректен"])

        headers = [h.strip().lower() for h in rows[0]]

        if "категория" not in headers or "тип" not in headers or "показатель" not in headers:
            return ImportResponse(
                imported=0, skipped=0,
                errors=["Файл не является экспортом бюджета (отсутствуют колонки 'Категория', 'Тип' или 'Показатель')"]
            )

        cat_idx = headers.index("категория")
        type_idx = headers.index("тип")
        metric_idx = headers.index("показатель")

        week_cols = {}
        for idx, h in enumerate(headers):
            if h.startswith("нед"):
                try:
                    week_num = int(h.replace("нед", ""))
                    if 1 <= week_num <= 52:
                        week_cols[week_num - 1] = idx
                except ValueError:
                    continue

        for i, row in enumerate(rows[1:], start=2):
            try:
                if not row or len(row) < 3:
                    continue

                cat_name = row[cat_idx].strip()
                type_str = row[type_idx].strip().lower()
                metric_str = row[metric_idx].strip().lower()

                if not cat_name:
                    skipped += 1
                    continue

                db_type = "income" if type_str in ("доход", "income") else "expense"

                values = [0.0] * 52
                for week_idx, file_col_idx in week_cols.items():
                    if file_col_idx < len(row):
                        val_str = str(row[file_col_idx]).strip()
                        if val_str:
                            try:
                                values[week_idx] = float(val_str.replace(',', '.').replace(' ', ''))
                            except ValueError:
                                pass

                existing = db.query(BudgetCategory).filter(
                    BudgetCategory.name == cat_name,
                    BudgetCategory.type == db_type
                ).first()

                if existing:
                    if metric_str in ("план", "plan"):
                        existing.monthly = json.dumps(values)
                    elif metric_str in ("факт", "fact"):
                        existing.monthly_fact = json.dumps(values)
                else:
                    new_monthly = json.dumps(values) if metric_str in ("план", "plan") else json.dumps([0.0] * 52)
                    new_monthly_fact = json.dumps(values) if metric_str in ("факт", "fact") else json.dumps([0.0] * 52)

                    new_cat = BudgetCategory(
                        name=cat_name,
                        type=db_type,
                        monthly=new_monthly,
                        monthly_fact=new_monthly_fact,
                        sort_order=i,
                    )
                    db.add(new_cat)
                imported += 1

            except Exception as e:
                errors.append(f"Строка {i}: {str(e)}")
                skipped += 1

        db.commit()

    except Exception as e:
        errors.append(f"Ошибка при импорте CSV: {str(e)}")

    return ImportResponse(imported=imported, skipped=skipped, errors=errors)


@app.post("/api/import/excel", response_model=ImportResponse)
async def import_excel(
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
):
    """Import budget table from Excel."""
    imported = 0
    skipped = 0
    errors: list[str] = []

    try:
        content = await file.read()
        with load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True) as wb:
            ws = wb.active
            if ws is None:
                return ImportResponse(imported=0, skipped=0, errors=["Не удалось открыть лист Excel"])

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return ImportResponse(imported=0, skipped=0, errors=["Файл пуст или содержит мало данных"])

            headers = []
            header_idx = -1

            for r_idx in range(min(10, len(rows))):
                row_cells = [str(c).strip().lower() if c is not None else "" for c in rows[r_idx]]
                if "категория" in row_cells and "тип" in row_cells and "показатель" in row_cells:
                    headers = row_cells
                    header_idx = r_idx
                    break

            if header_idx == -1:
                return ImportResponse(imported=0, skipped=0, errors=[
                    "Не удалось найти строку заголовков с обязательными столбцами 'Категория', 'Тип' и 'Показатель'"
                ])

            cat_idx = headers.index("категория")
            type_idx = headers.index("тип")
            metric_idx = headers.index("показатель")

            week_cols = {}
            for idx, h in enumerate(headers):
                if h.startswith("нед"):
                    try:
                        week_num = int(h.replace("нед", ""))
                        if 1 <= week_num <= 52:
                            week_cols[week_num - 1] = idx
                    except ValueError:
                        continue

            for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                try:
                    if not any(cell is not None for cell in row):
                        continue

                    def get_cell_str(idx):
                        return str(row[idx]).strip() if idx is not None and idx < len(row) and row[
                            idx] is not None else ""

                    cat_name = get_cell_str(cat_idx)
                    type_str = get_cell_str(type_idx).lower()
                    metric_str = get_cell_str(metric_idx).lower()

                    if not cat_name:
                        skipped += 1
                        continue

                    db_type = "income" if type_str in ("доход", "income") else "expense"

                    values = [0.0] * 52
                    for week_idx, file_col_idx in week_cols.items():
                        val_str = get_cell_str(file_col_idx)
                        if val_str:
                            try:
                                values[week_idx] = float(val_str.replace(',', '.').replace(' ', ''))
                            except ValueError:
                                pass

                    existing = db.query(BudgetCategory).filter(
                        BudgetCategory.name == cat_name,
                        BudgetCategory.type == db_type
                    ).first()

                    if existing:
                        if metric_str in ("план", "plan"):
                            existing.monthly = json.dumps(values)
                        elif metric_str in ("факт", "fact"):
                            existing.monthly_fact = json.dumps(values)
                    else:
                        new_monthly = json.dumps(values) if metric_str in ("план", "plan") else json.dumps([0.0] * 52)
                        new_monthly_fact = json.dumps(values) if metric_str in ("факт", "fact") else json.dumps(
                            [0.0] * 52)

                        new_cat = BudgetCategory(
                            name=cat_name,
                            type=db_type,
                            monthly=new_monthly,
                            monthly_fact=new_monthly_fact,
                            sort_order=i,
                        )
                        db.add(new_cat)
                    imported += 1

                except Exception as e:
                    errors.append(f"Строка {i}: {str(e)}")
                    skipped += 1

            db.commit()

    except Exception as e:
        errors.append(f"Ошибка при импорте Excel: {str(e)}")

    return ImportResponse(imported=imported, skipped=skipped, errors=errors)


@app.post("/api/import/sber-pdf", response_model=ImportResponse)
async def import_sber_pdf(
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
):
    """Import transactions directly from Sberbank PDF statements."""
    imported = 0
    skipped = 0
    errors: list[str] = []

    try:
        content = await file.read()
        pdf_file = io.BytesIO(content)
        reader = PdfReader(pdf_file)

        tx_header_re = re.compile(
            r"^(\d{2}\.\d{2}\.\d{4})\s+(\d{2}:\d{2})\s+(.+?)\s+([+-]?\s*[\d\s]+,\d{2})\s+([\d\s]+,\d{2})$"
        )

        tx_detail_re = re.compile(
            r"^(\d{2}\.\d{2}\.\d{4})\s+(\d{6})\s+(.+)$"
        )

        parsed_transactions = []
        current_tx = None

        for page in reader.pages:
            text = page.extract_text()
            if not text:
                continue

            lines = text.split("\n")
            for line in lines:
                line = line.strip()

                match_header = tx_header_re.match(line)
                if match_header:
                    if current_tx:
                        parsed_transactions.append(current_tx)

                    date_str, time_str, category, amount_str, _ = match_header.groups()
                    current_tx = {
                        "date": date_str,
                        "time": time_str,
                        "category": category.strip(),
                        "amount_str": amount_str.strip(),
                        "description": "",
                    }
                    continue

                match_detail = tx_detail_re.match(line)
                if match_detail:
                    date_str, _, description = match_detail.groups()
                    if current_tx and current_tx["date"] == date_str:
                        current_tx["description"] = description.strip()
                        parsed_transactions.append(current_tx)
                        current_tx = None
                    continue

        if current_tx:
            parsed_transactions.append(current_tx)

        for tx_data in parsed_transactions:
            try:
                day, month, year = tx_data["date"].split(".")
                iso_date = f"{year}-{month}-{day}"

                clean_amount_str = tx_data["amount_str"].replace("\xa0", "").replace(" ", "").replace(",", ".")

                if clean_amount_str.startswith("+"):
                    tx_type = "income"
                    amount = float(clean_amount_str.replace("+", ""))
                else:
                    tx_type = "expense"
                    amount = float(clean_amount_str)

                exists = db.query(Transaction).filter(
                    Transaction.date == iso_date,
                    Transaction.amount == amount,
                    Transaction.category == tx_data["category"],
                    Transaction.description == tx_data["description"]
                ).first()

                if exists:
                    skipped += 1
                    continue

                new_tx = Transaction(
                    date=iso_date,
                    category=tx_data["category"],
                    description=tx_data["description"] or f"Операция {tx_data['category']}",
                    amount=amount,
                    type=tx_type,
                )
                db.add(new_tx)
                imported += 1

            except Exception as row_err:
                errors.append(f"Ошибка при обработке строки за {tx_data.get('date')}: {str(row_err)}")
                skipped += 1

        db.commit()

    except Exception as e:
        errors.append(f"Не удалось обработать PDF-файл: {str(e)}")

    return ImportResponse(imported=imported, skipped=skipped, errors=errors)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  EXPORT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.get("/api/export/csv")
def export_csv(db: Session = Depends(get_db)):
    """Export budget table as CSV (UTF-8 with BOM for Excel compatibility)."""
    cats = db.query(BudgetCategory).order_by(BudgetCategory.sort_order).all()
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    headers = ["Категория", "Тип", "Показатель"] + [f"Нед{i + 1}" for i in range(52)]
    writer.writerow(headers)

    for cat in cats:
        type_label = "Доход" if cat.type == "income" else "Расход"

        monthly_plan = json.loads(cat.monthly) if cat.monthly else [0.0] * 52
        while len(monthly_plan) < 52:
            monthly_plan.append(0.0)

        monthly_fact = json.loads(cat.monthly_fact) if cat.monthly_fact else [0.0] * 52
        while len(monthly_fact) < 52:
            monthly_fact.append(0.0)

        writer.writerow([cat.name, type_label, "План"] + monthly_plan)
        writer.writerow([cat.name, type_label, "Факт"] + monthly_fact)

    csv_text = output.getvalue()
    bom = b'\xef\xbb\xbf'
    csv_bytes = bom + csv_text.encode('utf-8')

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=budget.csv"},
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  RESET
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


@app.post("/api/reset", status_code=200)
def reset_data(db: Session = Depends(get_db)):
    """Reset all data — drop and recreate all tables, then re-seed."""
    db.close()
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    seed_data()
    logger.info("All data has been reset")
    return {"message": "Все данные сброшены"}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


def _cat_to_dict(cat: BudgetCategory) -> dict:
    """Convert a SQLAlchemy BudgetCategory to a response dict."""
    monthly = json.loads(cat.monthly)
    monthly_fact = json.loads(cat.monthly_fact) if cat.monthly_fact else [0.0] * 52
    # Ensure 52 elements
    while len(monthly) < 52:
        monthly.append(0)
    while len(monthly_fact) < 52:
        monthly_fact.append(0)
    return {
        "id": cat.id,
        "name": cat.name,
        "type": cat.type,
        "monthly": monthly,
        "monthly_fact": monthly_fact,
        "sort_order": cat.sort_order,
        "parent_id": cat.parent_id,
        "created_at": cat.created_at,
    }


def _get_weekly_dates(db: Session) -> list[str]:
    """Compute 52 weekly start dates from the budgetStartDate setting."""
    setting = db.query(Setting).filter(Setting.key == "budgetStartDate").first()
    start_date_str = setting.value if setting and setting.value else ""

    if start_date_str:
        try:
            start = datetime.strptime(start_date_str, "%Y-%m-%d")
        except ValueError:
            start = datetime.now()
    else:
        # Default to Monday of current week
        today = datetime.now()
        start = today - timedelta(days=today.weekday())

    dates = []
    for i in range(52):
        d = start + timedelta(weeks=i)
        dates.append(d.strftime("%Y-%m-%d"))
    return dates


def _build_budget_response(
    cats: list[BudgetCategory],
    scenario: Scenario | None = None,
    db: Session = None,
) -> dict:
    """Build the full budget table response with optional scenario modifiers."""
    cat_responses = []
    monthly_income = [0.0] * 52
    monthly_expense = [0.0] * 52
    monthly_fact_income = [0.0] * 52
    monthly_fact_expense = [0.0] * 52

    for cat in cats:
        monthly = json.loads(cat.monthly)
        monthly_fact = json.loads(cat.monthly_fact) if cat.monthly_fact else [0.0] * 52

        # Ensure 52 elements
        while len(monthly) < 52:
            monthly.append(0)
        while len(monthly_fact) < 52:
            monthly_fact.append(0)

        # Apply scenario modifiers (only to plan, not fact)
        if scenario:
            if cat.type == "income":
                monthly = [round(v * scenario.income_modifier, 2) for v in monthly]
            else:
                per_week_extra = scenario.extra_expense / 52
                monthly = [round(v * scenario.expense_modifier + per_week_extra, 2) for v in monthly]

        cat_responses.append({
            "id": cat.id,
            "name": cat.name,
            "type": cat.type,
            "monthly": monthly,
            "monthly_fact": monthly_fact,
            "sort_order": cat.sort_order,
            "parent_id": cat.parent_id,
            "created_at": cat.created_at,
        })

        for i in range(52):
            if cat.type == "income":
                monthly_income[i] += monthly[i]
                monthly_fact_income[i] += monthly_fact[i]
            else:
                monthly_expense[i] += monthly[i]
                monthly_fact_expense[i] += monthly_fact[i]

    monthly_balance = [monthly_income[i] - monthly_expense[i] for i in range(52)]
    monthly_fact_balance = [monthly_fact_income[i] - monthly_fact_expense[i] for i in range(52)]

    weekly_remainder = [0.0] * 52
    weekly_wallet_total = [0.0] * 52
    weekly_cumulative_balance = [0.0] * 52
    
    weekly_fact_remainder = [0.0] * 52
    weekly_fact_wallet_total = [0.0] * 52
    weekly_fact_cumulative_balance = [0.0] * 52

    # Calculate cumulative balances for Plan
    curr_rem = 0.0
    for i in range(52):
        weekly_remainder[i] = curr_rem
        weekly_wallet_total[i] = curr_rem + monthly_income[i]
        weekly_cumulative_balance[i] = weekly_wallet_total[i] - monthly_expense[i]
        curr_rem = weekly_cumulative_balance[i]

    # Calculate cumulative balances for Fact
    curr_fact_rem = 0.0
    for i in range(52):
        weekly_fact_remainder[i] = curr_fact_rem
        weekly_fact_wallet_total[i] = curr_fact_rem + monthly_fact_income[i]
        weekly_fact_cumulative_balance[i] = weekly_fact_wallet_total[i] - monthly_fact_expense[i]
        curr_fact_rem = weekly_fact_cumulative_balance[i]

    # Total profit = sum of all weekly balances (unchanged logic to keep compatibility, or we could use the final cumulative balance)
    total_profit = sum(monthly_balance)

    # Weekly dates
    weekly_dates = _get_weekly_dates(db) if db else [f"2026-W{i+1:02d}" for i in range(52)]

    return {
        "categories": cat_responses,
        "monthly_income": monthly_income,
        "monthly_expense": monthly_expense,
        "monthly_balance": monthly_balance,
        "monthly_fact_income": monthly_fact_income,
        "monthly_fact_expense": monthly_fact_expense,
        "monthly_fact_balance": monthly_fact_balance,
        "weekly_dates": weekly_dates,
        "total_profit": total_profit,
        "weekly_remainder": weekly_remainder,
        "weekly_wallet_total": weekly_wallet_total,
        "weekly_cumulative_balance": weekly_cumulative_balance,
        "weekly_fact_remainder": weekly_fact_remainder,
        "weekly_fact_wallet_total": weekly_fact_wallet_total,
        "weekly_fact_cumulative_balance": weekly_fact_cumulative_balance,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  FRONTEND STATIC FILES (must be LAST — catch-all)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


if os.path.isdir(DIST_DIR):
    # Mount assets sub-directory for JS/CSS
    assets_dir = os.path.join(DIST_DIR, "assets")
    if os.path.isdir(assets_dir):
        app.mount("/assets", StaticFiles(directory=assets_dir), name="assets")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        """Serve static files or fallback to index.html for SPA routing."""
        file_path = os.path.join(DIST_DIR, full_path)
        if full_path and os.path.isfile(file_path):
            return FileResponse(file_path)
        index_path = os.path.join(DIST_DIR, "index.html")
        if os.path.isfile(index_path):
            return FileResponse(index_path)
        return HTMLResponse("<h1>Frontend not found</h1>", status_code=404)
